# -*- coding: utf-8 -*-
"""
Created on Sat May  5 15:56:56 2018

@author: Han
"""

import numpy as np
from openpyxl import Workbook

def nparray2string(val):
    '''
    convert a numpy array to string
    '''
    x = np.array2string(val, threshold=np.nan,
                                 max_line_width=np.inf,
                                 formatter={'int': lambda x: '%d' % (x),
                                            'float_kind':lambda x: "%.6e" % x}).replace('[', '').replace(']', '') + '\n'
    if val.ndim == 1:
        return x
    else:
        return ' ' + x

def py2dat(tdata,fname):
    if not isinstance(tdata,dict):
        raise TypeError
    else:
        with open(fname,'w',encoding='utf-8') as f:
            # variables
            f.write('VARIABLES = ')
            for var in tdata['varnames'][:-1]:
                f.write('"%s", '%(var,))
            f.write('"%s"\n'%(tdata['varnames'][-1],))

            nzone = -1

            # write lines
            if 'lines' in tdata.keys():
                for iline,line in enumerate(tdata['lines']):
                    nzone += 1
                    x = np.asarray(line['x'])
                    lenx = x.size
                    data = x.reshape(1,lenx)

                    y = np.asarray(line['y']).reshape(1,lenx)
                    data = np.vstack((data,y))

                    if 'z' in line.keys():
                        if line['z'].size > 0:
                            z = np.asarray(line['z']).reshape(1,lenx)
                            data = np.vstack((data,z))
                    if 'v' in line.keys():
                        if line['v'].size > 0:
                            v = np.asarray(line['v'])
                            if v.ndim == 1:
                                v = v.reshape(1,lenx)
                            data = np.vstack((data,v))

                    if 'zonename' in line.keys():
                        if len(line['zonename']) == 0:
                            zonename = 'ZONE %d'%(iline,)
                        else:
                            zonename = line['zonename']
                    else:
                        zonename = 'ZONE %d'%(nzone,)

                    ivarloc = 0
                    f.write('ZONE I = %d T="%s" DATAPACKING=POINT\n'%(lenx,zonename))
                    f.write(' '+np.array2string(data.T,threshold=np.nan,max_line_width=np.inf).replace('[','').replace(']','') )
                    f.write('\n\n')

            # write surfaces
            if 'surfaces' in tdata.keys():
                for isurf,surf in enumerate(tdata['surfaces']):
                    nzone += 1
                    # 0 for point, 1 for block
                    if 'datapacking' in surf.keys():
                        ipack = 'POINT' if surf['datapacking'] == 0 else 'BLOCK'
                    else:
                        ipack = 'POINT'

                    # 0 for nodal, 1 for center
                    if 'varloc' in surf.keys():
                        ivarloc = surf['varloc']
                        if isinstance(ivarloc, list):
                            ipack = 'BLOCK'
                            icen = []
                            inodal = []
                            for i, ii in enumerate(ivarloc):
                                if ii == 1:
                                    icen.append(i+1)
                                else:
                                    inodal.append(i+1)
                    else:
                        ivarloc = 0


                    # 3 for IJ order, 2 for IK order, 1 for JK order
                    if 'order' in surf.keys():
                        iorder = surf['order']
                    else:
                        iorder = 3


                    x = surf['x']
                    # x should be store in the following way
                    # x -----> i
                    # |
                    # |
                    # ^ j
                    y = surf['y']
                    if 'z' in surf.keys():
                        z = surf['z']
                    if 'v' in surf.keys():
                        v = surf['v']
                    if 'zonename' in surf.keys():
                        if len(surf['zonename']) == 0:
                            zonename = 'ZONE %d'%(nzone,)
                        else:
                            zonename = surf['zonename']

                    m, n = x.shape
                    f.write('ZONE I=%d, J=%d, T="%s", DATAPACKING=%s, '%(m,n,zonename,ipack))
                    if isinstance(ivarloc, list):
                        f.write('VARLOCATION=(%s=CELLCENTERED, %s=NODAL)\n'%(str(icen), str(inodal)))
                    elif ivarloc == 1:
                        f.write('VARLOCATION=([%d-%d]=CELLCENTERED)\n'%(1, len(tdata['varnames'])))

                    if ipack == 'BLOCK':
                        f.write(nparray2string(x.flatten()))
                        f.write(nparray2string(y.flatten()))
                        if 'z' in surf.keys():
                            f.write(nparray2string(z.flatten()))
                        if 'v' in surf.keys():
                            for vv in v:
                                f.write(nparray2string(vv.flatten()))
                    else:
                        data = x.flatten()
                        data = np.vstack((data,y.flatten()))
                        if 'z' in surf.keys():
                            data = np.vstack((data,z.flatten()))
                        if 'v' in line.keys():
                            for vv in v:
                                data = np.vstack((data,vv.flatten()))
                        f.write(nparray2string(data.T))
                    f.write('\n\n')

def py2xls(tdata, fname, creator='Han Luo'):
    wb = Workbook()
    k =-1
    for line in tdata['lines']:
        k = k+1
        if 'zonename' in line.keys():
            if len(line['zonename']) == 0:
                zn = 'Sheet %d'%(k+1,)
            else:
                zn = line['zonename']
        else:
            zn = 'Sheet %d'%(k+1,)

        ws=wb.create_sheet(zn,k)
        for i,j in enumerate(tdata['varnames']):
            cellnum = chr(65+i)
            ws[cellnum+'1'] = j
        data = np.vstack((line['x'],line['y'],line['z'],line['v'])).T
        for i,row in enumerate(data):
            for j,ele in enumerate(row):
                cellnum = chr(65+j)
                cellid = cellnum+'%d'%(i+2,)
                ws[cellid] = ele

    wb.properties.creator = creator
    wb.save(fname)

def py2dat2(tdata, fname):
    with open(fname, 'w') as f:
        # variables
        f.write('VARIABLES = ')
        for var in tdata['varnames'][:-1]:
            f.write('"%s", '%(var,))
        f.write('"%s"\n'%(tdata['varnames'][-1],))

        # data
        for zone in tdata['data']:
            f.write('\n')
            m, n = zone['data'].shape
            f.write('ZONE I = %d, T = "%s"\n'%(m, zone['title']))
            if 'passivevarlist' in zone.keys():
                if zone['passivevarlist']:
                    f.write('PASSIVEVARLIST=%s\n'%(str(zone['passivevarlist']),))
            f.write(nparray2string(zone['data']))


